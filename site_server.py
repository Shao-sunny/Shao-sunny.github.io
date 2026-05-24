import json
import os
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from socketserver import ThreadingMixIn
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "website_texts.xlsx"


def read_table(workbook, sheet_name):
    sheet = workbook[sheet_name]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    headers = [str(value) if value is not None else "" for value in rows[0]]
    records = []

    for values in rows[1:]:
      if not any(value not in (None, "") for value in values):
        continue
      record = {}
      for index, header in enumerate(headers):
        record[header] = values[index] if index < len(values) else ""
      records.append(record)

    return records


def workbook_to_data():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    settings = {row["key"]: row["value"] for row in read_table(workbook, "Settings")}
    home_lines = read_table(workbook, "HomeLines")
    about_info = read_table(workbook, "AboutInfo")
    education = read_table(workbook, "Education")
    education_details = read_table(workbook, "EducationDetails")
    internship_settings = {row["key"]: row["value"] for row in read_table(workbook, "InternshipSettings")}
    internship_details = read_table(workbook, "InternshipDetails")
    internship_results = read_table(workbook, "InternshipResults")
    sop_sections = read_table(workbook, "SOPSections")
    projects = read_table(workbook, "Projects")
    project_bullets = read_table(workbook, "ProjectBullets")
    videos = read_table(workbook, "Videos")
    posters = read_table(workbook, "Posters")
    wechat_rows = read_table(workbook, "Wechat")
    offline = read_table(workbook, "OfflineMaterials")

    education_detail_map = {}
    for item in education_details:
        education_detail_map.setdefault(item["item_id"], []).append(item)

    project_bullet_map = {}
    for item in project_bullets:
        project_bullet_map.setdefault(item["item_id"], []).append(item)

    sop_map = {}
    for item in sop_sections:
        section = sop_map.setdefault(
            item["section_id"],
            {
                "order": item["section_order"],
                "title": item["section_title"],
                "pills": [part for part in str(item["pill_group"] or "").split("|") if part],
                "items": [],
            },
        )
        section["items"].append((item["item_order"], item["item_text"]))

    education_items = []
    for item in sorted(education, key=lambda row: row["order"]):
        details = [detail["text"] for detail in sorted(education_detail_map.get(item["item_id"], []), key=lambda row: row["order"])]
        education_items.append(
            {
                "title": item["title"],
                "time": item["time"],
                "meta": item["meta"],
                "details": details,
            }
        )

    project_items = []
    for item in sorted(projects, key=lambda row: row["order"]):
        bullets = [bullet["text"] for bullet in sorted(project_bullet_map.get(item["item_id"], []), key=lambda row: row["order"])]
        project_items.append(
            {
                "title": item["title"],
                "role": item["role"],
                "time": item["time"],
                "description": item["description"],
                "bullets": bullets,
            }
        )

    sop_items = []
    for _, value in sorted(sop_map.items(), key=lambda pair: pair[1]["order"]):
        sop_items.append(
            {
                "title": value["title"],
                "pills": value["pills"],
                "items": [text for _, text in sorted(value["items"], key=lambda pair: pair[0])],
            }
        )

    return {
        "settings": {
            "site_title": settings.get("site_title", ""),
            "name": settings.get("name", ""),
            "home_eyebrow": settings.get("home_eyebrow", ""),
            "home_button_label": settings.get("home_button_label", ""),
            "home_button_target": settings.get("home_button_target", ""),
            "about_photo": settings.get("about_photo", ""),
            "poster_hint": settings.get("poster_hint", ""),
        },
        "home": {
            "cn_lines": [row["text"] for row in sorted(home_lines, key=lambda row: row["order"]) if row["lang"] == "cn"],
            "en_lines": [row["text"] for row in sorted(home_lines, key=lambda row: row["order"]) if row["lang"] == "en"],
        },
        "about": {
            "info": [{"label": row["label"], "value": row["value"]} for row in sorted(about_info, key=lambda row: row["order"])],
            "education": education_items,
        },
        "internship": {
            "company": internship_settings.get("company", ""),
            "role": internship_settings.get("role", ""),
            "time": internship_settings.get("time", ""),
            "overview_label": settings.get("internship_overview_label", "工作内容"),
            "sop_label": settings.get("internship_sop_label", "客户访谈视频SOP"),
            "details": [{"title": row["title"], "text": row["text"]} for row in sorted(internship_details, key=lambda row: row["order"])],
            "results": [
                {
                    "label": row["label"],
                    "title": row["title"],
                    "description": row["description"],
                    "action_type": row["action_type"],
                    "action_value": row["action_value"],
                    "action_tab": row["action_tab"],
                    "action_subtab_group": row["action_subtab_group"],
                    "action_subtab": row["action_subtab"],
                }
                for row in sorted(internship_results, key=lambda row: row["order"])
            ],
            "sop_intro": internship_settings.get("sop_intro", ""),
            "sop_sections": sop_items,
        },
        "projects": project_items,
        "works": {
            "video_label": settings.get("video_label", "视频作品"),
            "graphic_label": settings.get("graphic_label", "图文作品"),
            "poster_title": settings.get("poster_title", "海报"),
            "offline_title": settings.get("offline_title", "线下物料"),
            "videos": [{"title": row["title"], "file": row["file"]} for row in sorted(videos, key=lambda row: row["order"])],
            "posters": [{"title": row["title"], "file": row["file"]} for row in sorted(posters, key=lambda row: row["order"])],
            "wechat": wechat_rows[0] if wechat_rows else {"title": "", "note": "", "image": "", "url": ""},
            "offline": [{"title": row["title"], "file": row["file"]} for row in sorted(offline, key=lambda row: row["order"])],
        },
    }


class SiteHandler(SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/site-data":
            payload = json.dumps(workbook_to_data(), ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        super().do_GET()


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


def main():
    port = int(os.environ.get("SITE_PORT", "8765"))
    os.chdir(ROOT)
    server = ThreadingHTTPServer(("127.0.0.1", port), SiteHandler)
    print(f"http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
