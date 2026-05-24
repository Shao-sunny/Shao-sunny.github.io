from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "website_texts.xlsx"


def append_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    return sheet


def autofit(workbook):
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 4
            column_ref = column[0].column
            if isinstance(column_ref, int):
                column_letter = get_column_letter(column_ref)
            else:
                column_letter = column_ref
            sheet.column_dimensions[column_letter].width = min(width, 80)


def main():
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_sheet(
        workbook,
        "Settings",
        ["key", "value"],
        [
            ["site_title", "邵帅｜个人主页"],
            ["name", "邵帅"],
            ["home_eyebrow", "Personal Website"],
            ["home_button_label", "进入关于我"],
            ["home_button_target", "about"],
            ["about_photo", "邵帅照片.jpg"],
            ["logo_file", "Logo-clean.png"],
            ["poster_hint", "右滑查看更多"],
            ["video_label", "视频作品"],
            ["graphic_label", "图文作品"],
            ["poster_title", "海报"],
            ["wechat_title", "公众号推送"],
            ["wechat_note", "点击图片跳转至推文"],
            ["offline_title", "线下物料"],
            ["internship_overview_label", "工作内容"],
            ["internship_sop_label", "客户访谈视频SOP"],
        ],
    )

    append_sheet(
        workbook,
        "HomeLines",
        ["order", "lang", "text"],
        [
            [1, "cn", "用影像与设计勾勒灵感，"],
            [2, "cn", "用数据与新兴技术定义方向。"],
            [3, "cn", "在感性与理性之间，探索内容和产品的温度。"],
            [4, "en", "Create with taste."],
            [5, "en", "Think with data."],
            [6, "en", "Build with emerging tools."],
        ],
    )

    append_sheet(
        workbook,
        "AboutInfo",
        ["order", "label", "value"],
        [
            [1, "姓名", "邵帅"],
            [2, "手机", "18257100595"],
            [3, "邮箱", "1509341328@qq.com"],
            [4, "出生年月", "2005年3月2日"],
            [5, "现居地", "上海 / 杭州"],
        ],
    )

    append_sheet(
        workbook,
        "Education",
        ["item_id", "order", "title", "time", "meta"],
        [
            ["edu_1", 1, "杭州外国语学校（中学）", "2020.9 - 2023.7", ""],
            ["edu_2", 2, "上海交通大学（本科）", "2023.9 - 2027.6", "外国语学院 ｜ 日语专业（语言学方向）"],
            ["edu_3", 3, "早稻田大学（交换）", "2025.9 - 2026.2", "商学院"],
        ],
    )

    append_sheet(
        workbook,
        "EducationDetails",
        ["item_id", "order", "text"],
        [
            ["edu_2", 1, "成绩：综合学积分 91，GPA 3.9 / 4.3"],
            ["edu_2", 2, "荣誉与任职：曾获校学业优秀奖学金，任学院学生会宣传部部长、校学联干事、班级班长"],
            ["edu_2", 3, "相关课程：日语视听说（96）；日语精读（90）；语言数据与 Python 应用（86）；语言统计（95）；认知心理学（90）；影视制作实践（92）"],
            ["edu_3", 1, "相关课程：数字市场营销；市场管理；金融管理；国际贸易保险；公司并购与治理"],
        ],
    )

    append_sheet(
        workbook,
        "InternshipSettings",
        ["key", "value"],
        [
            ["company", "浙江有鹿机器人科技有限公司"],
            ["role", "市场部实习生（新媒体方向）"],
            ["time", "2025.6 - 2025.9"],
            ["sop_intro", "基于《有鹿客户采访系列视频 SOP》整理，围绕前期准备、拍摄执行与成片规范三个阶段，梳理出一套更适合网页展示的流程摘要。"],
        ],
    )

    append_sheet(
        workbook,
        "InternshipDetails",
        ["order", "title", "text"],
        [
            [1, "产品宣传与内容创作", "策划、制作、剪辑客户访谈系列视频，完成 5 条长视频，制定该系列视频 SOP，目前该系列在微信单平台获约 5000 曝光度。撰写脚本并制作双语版产品宣传视频 1 条，企业文字宣传材料海外化 2 份。"],
            [2, "调研与信息整合", "调研、沟通、比对出海营销投放、危机公关各 8+ 代理商服务与价格，汇总 2 份信息表供公司决策。"],
            [3, "商务工作", "全流程对接跟进服贸会展会工作，于省商务厅会议中独立进行企业及产品介绍展示；接待省市领导来访及媒体团队。"],
        ],
    )

    append_sheet(
        workbook,
        "InternshipResults",
        ["order", "label", "title", "description", "action_type", "action_value", "action_tab", "action_subtab_group", "action_subtab"],
        [
            [1, "成果 01", "客户访谈视频", "跳转至微信文章，查看实际客户访谈视频案例。", "external", "https://mp.weixin.qq.com/s/xOcf6KK07yETCq1gBSYdMQ", "", "", ""],
            [2, "成果 02", "客户访谈视频 SOP", "查看采访策划、拍摄流程与视频包装规范。", "internal", "", "internship", "internship", "internship-sop"],
            [3, "成果 03", "爆款视频分析", "打开 PDF 文件，查看视频分析成果。", "file", "爆款视频分析.pdf", "", "", ""],
        ],
    )

    append_sheet(
        workbook,
        "SOPSections",
        ["section_id", "section_order", "section_title", "pill_group", "item_order", "item_text"],
        [
            ["sop_1", 1, "前期工作", "", 1, "提前 5 至 7 日撰写采访提纲，先通过公开信息了解客户主体，判断其关注的价值点与决策角度。"],
            ["sop_1", 1, "前期工作", "", 2, "提纲建议包含采访目的、视频市场、采访对象、时间地点、模拟关键词和采访问题。"],
            ["sop_1", 1, "前期工作", "", 3, "固定问题聚焦引入机器人初衷、为何选择 AI130、使用后带来的最大改变，以及 3 个关键词总结产品价值。"],
            ["sop_1", 1, "前期工作", "", 4, "确认后导出 PDF，右上角加 UDEER·AI logo，并提前 2 至 3 日发送给被采访方准备。"],
            ["sop_1", 1, "前期工作", "", 5, "提前沟通摄像人员与设备，重点确认相机、麦克风、稳定器、三脚架、监听耳机等是否齐全且连接正常。"],
            ["sop_2", 2, "拍摄中期", "", 1, "采访总时长建议控制在 15 至 45 分钟，内容最好能够自然拆分为至少 3 个清晰要点。"],
            ["sop_2", 2, "拍摄中期", "", 2, "可以围绕现场反馈灵活追问，不必完全拘泥于原始提纲。"],
            ["sop_2", 2, "拍摄中期", "", 3, "正式拍摄前必须检查设备状态，避免镜头中出现故障、缺件、屏幕异常或明显脏污。"],
            ["sop_2", 2, "拍摄中期", "", 4, "场景镜头要尽量多样化，覆盖全景、中景、特写，以及正面、侧面、背面等不同角度。"],
            ["sop_2", 2, "拍摄中期", "", 5, "重点记录设备运行、贴边、避障、会车让车、面板或手机发任务等功能表现，并注意补充与采访内容匹配的画面。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 1, "封面以“客户说”为主标题，可搭配一句核心判断，例如“具身机器人重构服务价值”，使用渐显方式入场。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 2, "片头建议使用 4 至 5 个快切镜头，每个约 1 秒，形成更有节奏感的蒙太奇开场。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 3, "地点标注使用白色图标加“城市-地点名称”的方式呈现，并用轻提示音加强识别。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 4, "章节小标题可叠加 3 至 4 秒空镜，配合描边、阴影与滑动动画，强化段落切换的仪式感。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 5, "人名框采用半透明文本框与白色横线组合，姓名、单位和岗位分层呈现，增强品牌统一度。"],
            ["sop_3", 3, "视频内容规范", "封面及开头标题|片头快切|地点标注|章节小标题|人名框与关键词|字幕与片尾", 6, "字幕保持简洁易读，片尾可使用空镜加文字或纯黑底黄色字两种方案，并最终过渡到黑底黄 logo 页。"],
        ],
    )

    append_sheet(
        workbook,
        "Projects",
        ["item_id", "order", "title", "role", "time", "description"],
        [
            ["project_1", 1, "“星幻·Fantasia”年末 GALA", "总制作人", "2024.7 - 2024.12", "项目内容：排演 4 部经典音乐剧片段，并收入了 10 个独唱或合唱歌曲节目。邀请到 10+ 名音乐剧现役演员录制祝福作为预热视频。独立采购搭建追光、音控台、二层舞台等设备以呈现更好舞台效果。"],
            ["project_2", 2, "大学生创新创业训练计划", "", "2023.10 - 2025.4", ""],
            ["project_3", 3, "“中国种子”创新创业大赛", "首席运营官", "2022.6 - 2022.7", ""],
            ["project_4", 4, "AIGC助力中国影视作品出海的实践研究", "组长", "2024.8 - 2025.3", ""],
        ],
    )

    append_sheet(
        workbook,
        "ProjectBullets",
        ["item_id", "order", "text"],
        [
            ["project_1", 1, "【跨部门协调组织】通过飞书领导活动策划、执行的全流程工作，积累了项目运营经验。协调 4+ 部门近百名演职人员，提升了跨部门管理能力。对接、交涉 6+ 方企业及组织等，沟通版权问题、拉赞助、场地协调等，提升了外联能力和谈判技巧。"],
            ["project_1", 2, "【项目成果】拉动 8000 元校外赞助支持，总预算近 2w。推送阅读量达 7000+，现场满座（400+），转化社员 40+ 人。"],
            ["project_2", 1, "【学术研究】精读英文文献，进行 30+ 分钟汇报，撰写与华山医院合作科普文章 3 篇。开展实验 30+ 次，参与数据整理工作。"],
            ["project_2", 2, "【宣传投放与联络】制作宣传招募海报，通过线上线下多渠道招募被试，对接联络，累计 50+ 人。制作项目总结视频。"],
            ["project_3", 1, "【市场分析】基于 SWOT、波特五力等进行调研、数据分析等，明确痛点和需求，细化产品内容。"],
            ["project_3", 2, "【商业模式分析】协调团队分工，分析盈利模式、未来发展等，制作商业画布，形成商业计划书。"],
            ["project_4", 1, "【产品测试与落地】搭建、调试 AI 语言模型生成 5 期电台节目，第一期播放量较往期内容提升 4 倍播放量。"],
            ["project_4", 2, "【AI赋能视频创作】使用 AIGC 工具辅助，完成“敦煌光影展”实地佛像背后的动态光影背景设计。"],
        ],
    )

    append_sheet(
        workbook,
        "Videos",
        ["order", "title", "file"],
        [
            [1, "Fake Forture - 反金融欺诈宣传视频", "《Fake Fortune》反诈视频.mp4"],
            [2, "Room 606 - 原创恐怖微电影", "Room 606.mp4"],
        ],
    )

    append_sheet(
        workbook,
        "Posters",
        ["order", "title", "file"],
        [
            [1, "个人竞选海报", "图文作品/个人竞选海报.png"],
            [2, "讲座海报", "图文作品/讲座海报.jpg"],
            [3, "被试招募海报", "图文作品/被试招募海报.jpg"],
            [4, "节气海报", "图文作品/节气海报.jpg"],
        ],
    )

    append_sheet(
        workbook,
        "Wechat",
        ["title", "note", "image", "url"],
        [
            ["公众号推送", "点击图片跳转至推文", "图文作品/公众号封面.jpg", "https://mp.weixin.qq.com/s/Snl4LO-SHFck5RHLdULQgQ?snipping_id=0669c8858457a61995647c5a057b82fb"],
        ],
    )

    append_sheet(
        workbook,
        "OfflineMaterials",
        ["order", "title", "file"],
        [
            [1, "迎新手举牌", "图文作品/迎新手举牌.jpg"],
            [2, "迎新易拉宝", "图文作品/迎新易拉宝.jpg"],
        ],
    )

    autofit(workbook)
    workbook.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
