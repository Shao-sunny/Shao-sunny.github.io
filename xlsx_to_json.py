import json
from openpyxl import load_workbook

wb = load_workbook("website_texts.xlsx", data_only=True)

data = {}

for sheet in wb.sheetnames:
    ws = wb[sheet]

    rows = list(ws.values)

    if not rows:
        continue

    headers = rows[0]
    content = []

    for row in rows[1:]:
        item = {}

        for i, value in enumerate(row):
            key = headers[i]

            item[key] = value

        content.append(item)

    data[sheet] = content

with open("website_texts.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("已生成 website_texts.json")